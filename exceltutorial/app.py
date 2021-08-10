import openpyxl
from openpyxl import Workbook
import json
from flask import Flask

app = Flask(__name__)

@app.route('/')
def export_excel():

  return 'Hello, Export Excel!'

if __name__ == "__main__":
  # app.run(host ='0.0.0.0')
  json_data = {}

  with open("data.json") as json_file:
    json_data = json.load(json_file)
  
  # print(json_data)
  wb = Workbook()

  # Grab the active worksheet
  ws_01 = wb.active

  # Set the title of the worksheet
  ws_01.title = "First Sheet"

  # Set First row
  ws_01.cell(1,1, "Month")
  ws_01.cell(1,2, "Food")
  ws_01.cell(1,3, "Drink")
  ws_01.cell(1,4, "Price")

  row = 1
  for month in json_data.keys():
    row += 1
    ws_01.cell(row, 1 , month)
    ws_01.cell(row, 2 , json_data[month]["food"])
    ws_01.cell(row, 3 , json_data[month]["drink"])
    ws_01.cell(row, 4 , float(json_data[month]["price"]))

  # Save it in excel file
  wb.save("result_import.xlsx")
